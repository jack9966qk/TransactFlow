import csv
from datetime import datetime
from pathlib import Path

from .common import writeLocalTimeString
from .config import AmexJpRetrievalConfig
import xlrd
import openpyxl
import openpyxl.worksheet.worksheet

def moveFileForYearIntoDataDir(filePath: Path, name: str, config: AmexJpRetrievalConfig):
    moveToPath = config.yearsDir / f"{name}.xlsx"
    if moveToPath.exists(): moveToPath.unlink()
    filePath.rename(moveToPath)

def convertYearsXLSToCSV(config: AmexJpRetrievalConfig):
    # Remove existing converted files.
    for existing in config.convertedDir.iterdir():
        existing.unlink()

    for child in config.yearsDir.iterdir():
        name = child.stem
        ext = child.suffix
        def applyLineBreakEscape(s: str) -> str: return s.replace("\n", "\\n")
        if ext == ".xls":
            workbook = xlrd.open_workbook(str(child))
            assert(workbook.sheet_names()[0] =="ご利用金額")
            sheet = workbook.sheet_by_index(0)
            cellValues = [
                [applyLineBreakEscape(s) for s in sheet.row_values(r)]
                for r in range(sheet.nrows)
            ]
        elif ext == ".xlsx":
            workbook = openpyxl.load_workbook(filename=child, read_only=True)
            assert(workbook.sheetnames[0] == "ご利用履歴")
            sheet = workbook.worksheets[0]
            cellValues = [
                [applyLineBreakEscape("" if item is None else str(item)) for item in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        else:
            print(f"[FileMerge/AMEX JP] Skipping {child.name} under sheets directory")
            continue
        # There seems to be no indication of completeness in the file itself.
        yearComplete = int(name) < config.currentYear
        assert(datetime.now().year == config.currentYear)
        outputName = f"{name}.csv" if yearComplete else f"{name}_incomplete.csv"
        with open(config.convertedDir / outputName, "w") as outFile:
            csvWriter = csv.writer(outFile)
            for row in cellValues: csvWriter.writerow(row)

def updateFilesWithDownloadedXLSX(filePath: Path, name: str, config: AmexJpRetrievalConfig):
    moveFileForYearIntoDataDir(filePath, name, config)
    convertYearsXLSToCSV(config)
    writeLocalTimeString(config.timestampPath)
